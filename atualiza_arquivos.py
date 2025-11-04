
from hmac import new

from openpyxl.worksheet import worksheet
import data_eng 
import pandas as pd
import numpy as np
from openpyxl import load_workbook, workbook
import math
import requests
import datetime
import locale
from pprint import pprint

from tools.conab_dict import month_numeric, numeric_to_month

from tools.treatments_tools import list_is_equal

from make_support_data import make_relation_agrotoxics_cultures


def atualiza_VBP():
    new_values = data_eng.catch_vbp_values('./geral/VBP.xlsx')
    new_values = new_values['VBP']
    workbook = load_workbook('./dados_teste/VBP.xlsx')
    sheet = workbook.active 

    last_line = sheet.max_row
    print(sheet[last_line])

    ordem = ['ano', 'n', 'valor', 'variação (%)', 'mês', 'n mês', 'variação mês']
    valores = [new_values['ano'], sheet.max_row + 1, new_values['valor'], None, new_values['mes'], new_values['mes_num'], None]
    valid_data = False
    if range(len(sheet[last_line])) != range(len(valores)):
        raise ValueError("O tamanho de valores é diferente do tamanho da última linha da planilha")
    else:
        for i in range(len(sheet[last_line])):
            print(sheet[last_line][i].value, valores[i])
            if sheet[last_line][i].value != valores[i] and not ordem[i] == 'n':
                if ordem[i] == 'valor':
                    diference = valores[i] - sheet[last_line][i].value
                    print(diference)
                    if diference > 2:
                        print("A diferença é maior que 2")
                        valid_data = True
                    else:
                        print("Os valores são muito parecidos")
                else:
                    print("Os valores são iguais")
  
            else:
                print("Os valores são iguais")
            print('='*10)
        if valid_data:
            sheet.append(valores)
            workbook.save('./dados_teste/VBP.xlsx')

def atualiza_safras():
    cafe = data_eng.pega_valores_cafe()['cafe']
    cana = data_eng.pega_valores_cana()['cana']
    graos = data_eng.pega_valores_graos()['graos']


    cafe_m = [
        [cafe['ano'],float(f'{cafe['area'] /1000000:.2f}'),'Área(milhão ha)','Café'],
        [cafe['ano'],float(f'{cafe['produtividade']:.2f}'),'Produtividade(sc/há)','Café'],
        [cafe['ano'],float(f'{cafe['producao'] / 1000:.2f}'),'Produção(milhão sc)','Café'],
    ]
    cana_m = [
        [cana['ano'],float(f'{cana['area'] /1000:.2f}'),'Área(milhão ha)','Cana'],
        [cana['ano'],float(f'{cana['produtividade'] / 1000:.2f}'),'Produtividade(mil kg/há)','Cana'],
        [cana['ano'],float(f'{cana['producao'] / 1000:.2f}'),'Produção(milhão ton)','Cana'],
    ]
    graos_m = [
        [graos['ano'],float(f'{graos['area'] /1000:.2f}'),'Área(milhão ha)','Grão'],
        [graos['ano'],float(f'{graos['produtividade'] / 1000:.2f}'),'Produtividade(mil kg/há)','Grão'],
        [graos['ano'],float(f'{graos['producao'] / 1000:.2f}'),'Produção(milhão ton)','Grão'],
    ]
    workbook = load_workbook('./dados_teste/Safras.xlsx')
    sheet = workbook.active
    
    # Primeiro, remover todas as linhas de 2025 existentes
    print("Removendo dados de 2025 existentes...")
    rows_to_delete = []
    for i in range(1, sheet.max_row + 1):
        if sheet[i][0].value == '2025':
            rows_to_delete.append(i)
    
    # Deletar linhas de trás para frente para não afetar os índices
    for row_num in reversed(rows_to_delete):
        sheet.delete_rows(row_num)
        print(f"Removida linha {row_num} com dados de 2025")
    
    total_lines = sheet.max_row
    print(f"Total de linhas na planilha após remoção: {total_lines}")

    # Agora adicionar os novos dados de 2025
    for row_coffe in cafe_m:
        found_equal = False
        
        # Verificar todas as linhas (incluindo a última)
        for i in range(1, sheet.max_row + 1):
            row_planilha = [cell.value for cell in sheet[i]]
            if list_is_equal(row_coffe, row_planilha):
                print("Não adicionou: Havia linha igual")
                found_equal = True
                break
        
        # Só adiciona se NÃO encontrou nenhuma linha igual
        if not found_equal:
            print("Adicionando nova linha de café")
            sheet.append(row_coffe)
    
    for row_cana in cana_m:
        found_equal = False
        
        # Verificar todas as linhas (incluindo a última)
        for i in range(1, sheet.max_row + 1):
            row_planilha = [cell.value for cell in sheet[i]]
            if list_is_equal(row_cana, row_planilha):
                print("Não adicionou: Havia linha igual")
                found_equal = True
                break

        if not found_equal:
            print("Adicionando nova linha de cana")
            sheet.append(row_cana)

    for row_graos in graos_m:
        found_equal = False
        
        # Verificar todas as linhas (incluindo a última)
        for i in range(1, sheet.max_row + 1):
            row_planilha = [cell.value for cell in sheet[i]]
            if list_is_equal(row_graos, row_planilha):
                print("Não adicionou: Havia linha igual")
                found_equal = True
                break
                
        if not found_equal:
            print("Adicionando nova linha de grãos")
            sheet.append(row_graos)
    
    workbook.save('./dados_teste/Safras.xlsx')
    
def atualiza_dolar():
    def to_br_format(value):
        value = str(value)
        value = value.split('-')
        return f'{value[1]}-{value[2]}-{value[0]}'

   
    today = datetime.date.today()
    first_day = today.replace(day=1)
    
    today = to_br_format(today)
    first_day = to_br_format(first_day)

    url = f"https://olinda.bcb.gov.br/olinda/servico/PTAX/versao/v1/odata/CotacaoDolarPeriodo(dataInicial=@dataInicial,dataFinalCotacao=@dataFinalCotacao)?@dataInicial='{first_day}'&@dataFinalCotacao='{today}'&$top=100&$format=json&$select=cotacaoVenda"
    request = requests.get(url)
    data = request.json()
    valores = []
    for i in data['value']:
        valores.append(i['cotacaoVenda'])
    locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')
    data_to_append = [datetime.date.today().year,datetime.date.today().strftime('%b').capitalize(),f'{np.median(valores):.2f}', datetime.date.today().month]
    
    
    workbook = load_workbook('./dados_teste/Dolar.xlsx')
    sheet = workbook.active

    last_line = sheet.max_row

    last_line_sheet = [i.value for i in sheet[last_line]]
    print(last_line_sheet)

    if list_is_equal(last_line_sheet,data_to_append):
        print('Já atualizou')
        return
    sheet.append(data_to_append)
    workbook.save('./dados_teste/Dolar.xlsx')   

def atualiza_agrotoxicos():
    from datetime import datetime
    
    # Carrega dados de inseticidas e herbicidas
    inseticidas = data_eng.catch_values_conab('./geral/Agrotoxicos_Inseticida.xls', 'inseticidas')
    herbicidas = data_eng.catch_values_conab('./geral/Agrotoxicos_Herbicida.xls', 'herbicidas')
    fungicidas =  data_eng.catch_values_conab('./geral/Fertilizantes_Fungicida.xls', 'fungicidas')
    
    dado_apoio = make_relation_agrotoxics_cultures()
    # Carrega a planilha de destino
    workbook = load_workbook('./dados_teste/Agrotóxicos.xlsx')
    sheet = workbook.active
    
    # Ano atual
    ano_atual = datetime.now().year
    
    # Processa inseticidas
    for produto_dict in inseticidas:
        for produto_nome, dados in produto_dict.items():
            if dados['mediana'] and len(dados['mediana']) > 0:
                mediana_preco = dados['mediana'][0]
                unidade = dados['unidade'][0] if dados['unidade'] else 'L'
                
                # Procura o produto nos dados de apoio
                cultivo_encontrado = None
                preco_medio = None
                
                for item in dado_apoio:
                    for k, v in item.items():
                        if k != 'Preço Médio' and isinstance(v, list):
                            if produto_nome.upper() in v:
                                cultivo_encontrado = k
                                preco_medio = item['Preço Médio']
                                break
                    if cultivo_encontrado:
                        break
                
                # Cria linha de dados para adicionar
                nova_linha = [
                    ano_atual,           # Ano
                    'Inseticida',        # Sub-Grupo
                    produto_nome.upper(), # Produto
                    unidade,             # Unidade
                    mediana_preco,       # Preço (mediana)
                    cultivo_encontrado,  # Cultivo
                    preco_medio          # À vista R$/Kg
                ]
                
                # Verifica se já existe uma linha igual
                linha_existe = False
                for row in range(2, sheet.max_row + 1):  # Começa da linha 2 (pula cabeçalho)
                    linha_existente = [cell.value for cell in sheet[row]]
                    if list_is_equal(linha_existente, nova_linha):
                        linha_existe = True
                        break
                
                # Se não existe, adiciona ao final
                if not linha_existe:
                    sheet.append(nova_linha)
                    print(f'Adicionado inseticida: {produto_nome} - {mediana_preco} - Cultivo: {cultivo_encontrado}')
    
    # Processa herbicidas
    for produto_dict in herbicidas:
        for produto_nome, dados in produto_dict.items():
            if dados['mediana'] and len(dados['mediana']) > 0:
                mediana_preco = dados['mediana'][0]
                unidade = dados['unidade'][0] if dados['unidade'] else 'L'
                
                # Procura o produto nos dados de apoio
                cultivo_encontrado = None
                preco_medio = None
                
                for item in dado_apoio:
                    for k, v in item.items():
                        if k != 'Preço Médio' and isinstance(v, list):
                            if produto_nome.upper() in v:
                                cultivo_encontrado = k
                                preco_medio = item['Preço Médio']
                                break
                    if cultivo_encontrado:
                        break
                
                # Cria linha de dados para adicionar
                nova_linha = [
                    ano_atual,           # Ano
                    'Herbicida',         # Sub-Grupo
                    produto_nome.upper(), # Produto
                    unidade,             # Unidade
                    mediana_preco,       # Preço (mediana)
                    cultivo_encontrado,  # Cultivo
                    round(preco_medio,3)          # À vista R$/Kg
                ]
                
                # Verifica se já existe uma linha igual
                linha_existe = False
                for row in range(2, sheet.max_row + 1):  # Começa da linha 2 (pula cabeçalho)
                    linha_existente = [cell.value for cell in sheet[row]]
                    try:
                        linha_existente[-1] = round(linha_existente[-1], 3)
                    except (TypeError, ValueError):
                        pass
                    print("="*10)
                    print(linha_existente)
                    print(nova_linha)
                    if list_is_equal(linha_existente, nova_linha):
                        linha_existe = True
                        break
                
                # Se não existe, adiciona ao final
                if not linha_existe:
                    sheet.append(nova_linha)
                    print(f'Adicionado herbicida: {produto_nome} - {mediana_preco} - Cultivo: {cultivo_encontrado}')

    for produto_dict in fungicidas:
        for produto_nome, dados in produto_dict.items():
            if dados['mediana'] and len(dados['mediana']) > 0:
                mediana_preco = dados['mediana'][0]
                unidade = dados['unidade'][0] if dados['unidade'] else 'L'
                
                # Procura o produto nos dados de apoio
                cultivo_encontrado = None
                preco_medio = None
                
                for item in dado_apoio:
                    for k, v in item.items():
                        if k != 'Preço Médio' and isinstance(v, list):
                            if produto_nome.upper() in v:
                                cultivo_encontrado = k
                                preco_medio = item['Preço Médio']
                                break
                    if cultivo_encontrado:
                        break
                
                # Cria linha de dados para adicionar
                nova_linha = [
                    ano_atual,           # Ano
                    'Fungicida',        # Sub-Grupo
                    produto_nome.upper(), # Produto
                    unidade,             # Unidade
                    mediana_preco,       # Preço (mediana)
                    cultivo_encontrado,  # Cultivo
                    preco_medio          # À vista R$/Kg
                ]
                
                # Verifica se já existe uma linha igual
                linha_existe = False
                for row in range(2, sheet.max_row + 1):  # Começa da linha 2 (pula cabeçalho)
                    linha_existente = [cell.value for cell in sheet[row]]
                    if list_is_equal(linha_existente, nova_linha):
                        linha_existe = True
                        break
                
                # Se não existe, adiciona ao final
                if not linha_existe:
                    sheet.append(nova_linha)
                    print(f'Adicionado inseticida: {produto_nome} - {mediana_preco} - Cultivo: {cultivo_encontrado}')
    
    # Salva a planilha
    workbook.save('./dados_teste/Agrotóxicos.xlsx')
    print('Planilha de agrotóxicos atualizada com sucesso!')

def atualiza_credito_rural():
    df = data_eng.catch_credito_rural_values('./geral/Credito Rural.xlsx')
    
    df['Mes'] = df['Mês/Ano Protocolo'].apply(lambda x: numeric_to_month[int(x.split('/')[0].lstrip('0'))][0:3])
    df['Mes_numeric'] = df['Mês/Ano Protocolo'].apply(lambda x: x.split('/')[0])
    df['Ano'] = df['Mês/Ano Protocolo'].apply(lambda x: x.split('/')[-1])

    df.drop(columns={'Mês/Ano Protocolo'}, inplace=True)
    workbook = load_workbook('./dados_teste/Credito_Rural.xlsx')
    sheet = workbook.active

    max_rows = sheet.max_row

   
    for i, r in df.iterrows():
        x = None
        new_line = [
            r['Ano'],
            r['Valor comprometido R$'],
            None,
            r['Mes'],
            r['Mes_numeric'],
            x
        ]
    
        line_exists = False
        for row in sheet:
            row_correct = [cell.value for cell in row]
            
            

            if list_is_equal(row_correct, new_line):
                print(" Já Existia ", row_correct)
                line_exists = True
                break

        if not line_exists:
            x = round(((sheet[sheet.max_row][1].value - r['Valor comprometido R$']) / sheet[sheet.max_row][1].value) * 100, 3)
            print(f"-> {x}")
            new_line[-1] = x
            sheet.append(new_line)
            print(" Adicionou ", new_line)
    workbook.save('./dados_teste/Credito_Rural.xlsx')
    print(df)

def atualiza_indicadores_poda():
    podas = data_eng.catch_podas('./geral/Dados_para_poda.xlsx')
    workbook = load_workbook('./dados_teste/indicadores_Poda.xlsx')
    sheet = workbook.active

    max_rows = sheet.max_row
  
    for categoria, r in podas.iterrows():
        new_line = [
            categoria,
            datetime.datetime.now().year -1,
            float(r['Área colhida']),
            float(r['Quantidade produzida']),
            float(r['Custo por Hectare'])
        ]
      
        line_exists = False
        for row in sheet:
            row_correct = [i.value for i in row]
            if new_line[0] == row_correct[0] and new_line[1] == row_correct[1]:
                line_exists = True
                break            
        if not line_exists:
            print('ADICIONOU')
            sheet.append(new_line)
        
    workbook.save('./dados_teste/indicadores_Poda.xlsx')

def atualiza_indicadores_preços(): 
    try:
        # Carrega a planilha de preços de cultivos existente
        wb_destino = load_workbook('dados_teste/Preço Cultivos.xlsx')
        ws_destino = wb_destino.active
    except FileNotFoundError:
        # Se o arquivo não existir, cria um novo
        wb_destino = workbook.Workbook()
        ws_destino = wb_destino.active
        # Cria o cabeçalho
        ws_destino.append(['Produto', 'Nível de comercialização', 'UF', 'Mês', 'Ano', 'Preço medio(R$/Kg)'])

    # Carrega a planilha com os novos dados
    dados_novos = data_eng.catch_cultives_prices('geral/Preços Cultivos.xlsx')
    for novo_dado in dados_novos:
        nova_linha = [
            novo_dado['Produto'],
            novo_dado['Nivel de comercialização'],
            novo_dado['UF'],
            int(novo_dado['Mês'].lstrip('0')),
            int(novo_dado['Ano']),
            round(novo_dado['Preço Médio (R$/Kg)'], 2)
        ]

        # Verifica se a linha já existe na planilha de destino
        linha_existe = False
        for linha_destino in ws_destino.iter_rows(min_row=2, values_only=True):
            # Compara as colunas chave para identificar duplicatas
  
            if (nova_linha[0] == linha_destino[0] and      # Produto
                nova_linha[1] == linha_destino[1] and      # Nivel de comercialização
                nova_linha[2] == linha_destino[2] and      # UF
                nova_linha[3] == linha_destino[3] and      # Mês
                nova_linha[4] == linha_destino[4] and
                nova_linha[5] == linha_destino[5]):        # Ano
                linha_existe = True
                break
        
        # Se a linha não existir, adiciona à planilha de destino
        if not linha_existe:
            ws_destino.append(nova_linha)
            print(f"Linha adicionada: {nova_linha}")

    # Salva as alterações na planilha de destino
    wb_destino.save('dados_teste/Preço Cultivos.xlsx')
    print("Planilha 'Preço Cultivos.xlsx' atualizada com sucesso!")

def atualiza_pib_agro():
    pib = data_eng.catch_pib_prices('./geral/Pib Agro.pdf')

    print(pib)

    ano = datetime.datetime.now().year
    mes = datetime.datetime.now().month
    mes_str = numeric_to_month[mes][0:3]
    variation = None
    new_line = [ano, pib, mes, mes_str, variation]
    workbook = load_workbook('./dados_teste/PIB.xlsx')
    sheet = workbook.active
        # A nova linha a ser verificada e potencialmente adicionada
    # O valor da variação será preenchido após a verificação de existência
    new_line = [ano, pib, mes, mes_str, variation]
    print(f"Nova linha proposta: {new_line}")
    line_exists = False
    # Começa da segunda linha para ignorar o cabeçalho
    for row_idx in range(2, sheet.max_row + 1):
       row_correct = [cell.value for cell in sheet[row_idx]]
       # Compara apenas as colunas que identificam a singularidade (Ano, PIB, Mês, Mês_str)
       # Ignoramos a variação na comparação, pois ela pode ser diferente ou None
       if (new_line[0] == row_correct[0] and # Ano
           new_line[1] == row_correct[1] and # PIB
           new_line[2] == row_correct[2] and # Mês numérico
           new_line[3] == row_correct[3]):   # Mês strinhg
            line_exists = True
            print(f"Linha já existe: {row_correct}")
            break
        else:   
            sheet.append(new_line)
            print(f"Linha adicionada: {new_line}")
    else:
        print("Nenhuma nova linha adicionada, pois já existia uma correspondente.")

    workbook.save('./dados_teste/PIB.xlsx')
    print("Planilha 'Pib.xlsx' atualizada com sucesso!")
   
    
    
if __name__ == "__main__":
    # atualiza_credito_rural()
    # atualiza_agrotoxicos()
    # atualiza_safras()
    # atualiza_dolar()
    # atualiza_VBP()
    # atualiza_indicadores_preços()
    atualiza_pib_agro()
    